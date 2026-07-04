"""Excel reading using openpyxl (no pandas).

Reads the first worksheet, treats the first row as headers, and returns one
:class:`~src.models.Contact` per non-empty data row. Rows are positional —
the recipient is column A and the subject is built from the first columns.
"""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import load_workbook

from src.exceptions import EmptyWorkbookError, WorkbookError
from src.logger import get_logger
from src.models import Contact

log = get_logger(__name__)


class ExcelReader:
    """Reads contacts from an .xlsx workbook.

    Args:
        path:       Path to the workbook.
        sheet_name: Optional sheet to read; defaults to the first/active sheet.
    """

    def __init__(self, path: Path, sheet_name: str | None = None) -> None:
        self.path = Path(path)
        self.sheet_name = sheet_name

    def read(self) -> List[Contact]:
        """Read all non-empty contact rows from the first worksheet."""
        if not self.path.exists():
            raise WorkbookError(f"Workbook not found: {self.path}")

        rows_from_worksheet = self._read_rows_with_excel_com()
        if not rows_from_worksheet:
            rows_from_worksheet = self._read_rows_with_openpyxl()

        rows = iter(rows_from_worksheet)

        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise EmptyWorkbookError(f"Workbook is empty: {self.path}") from exc

        headers = [self._clean(c) for c in header_row]
        # Number of meaningful columns = up to the last non-empty header.
        col_count = self._last_non_empty_index(headers) + 1
        if col_count <= 0:
            raise EmptyWorkbookError(f"No header columns in: {self.path}")
        headers = headers[:col_count]

        contacts: List[Contact] = []
        for row_number, raw in enumerate(rows, start=2):
            if raw is None:
                continue

            values = [
                self._clean(raw[i] if i < len(raw) else None)
                for i in range(col_count)
            ]
            contact = Contact(
                values=values, row_number=row_number, headers=headers
            )
            if contact.is_empty():
                continue  # ignore empty rows
            contacts.append(contact)

        if not contacts:
            raise EmptyWorkbookError(
                f"No valid data rows found in workbook: {self.path}"
            )

        log.info("Read %d contact(s) from %s", len(contacts), self.path.name)
        return contacts

    def _read_rows_with_excel_com(self) -> List[List[object]]:
        """Read displayed cell text through Excel COM when available."""
        try:
            import win32com.client
        except ImportError:
            return []

        excel = None
        workbook = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(str(self.path), ReadOnly=True)
            worksheet = (
                workbook.Worksheets(self.sheet_name)
                if self.sheet_name
                else workbook.Worksheets(1)
            )
            used_range = worksheet.UsedRange
            row_count = used_range.Rows.Count
            col_count = used_range.Columns.Count

            rows: List[List[object]] = []
            for row_index in range(1, row_count + 1):
                row_values: List[object] = []
                for col_index in range(1, col_count + 1):
                    row_values.append(used_range.Cells(row_index, col_index).Text)
                rows.append(row_values)
            return rows
        except Exception as exc:
            log.warning("Excel COM read failed for %s: %s", self.path.name, exc)
            return []
        finally:
            if workbook is not None:
                workbook.Close(False)
            if excel is not None:
                excel.Quit()

    def _read_rows_with_openpyxl(self) -> List[List[object]]:
        """Fallback reader when Excel COM is unavailable."""
        wb = load_workbook(self.path, data_only=True)
        try:
            ws = wb[self.sheet_name] if self.sheet_name else wb.worksheets[0]
            return [
                [cell.value for cell in row]
                for row in ws.iter_rows()
            ]
        finally:
            wb.close()

    @staticmethod
    def _clean(value: object) -> str:
        """Trim a cell value to a string ('' for None)."""
        return "" if value is None else str(value).strip()

    @staticmethod
    def _last_non_empty_index(cells: List[str]) -> int:
        """Return the index of the last non-empty cell, or -1 if all empty."""
        for i in range(len(cells) - 1, -1, -1):
            if cells[i] != "":
                return i
        return -1


__all__ = ["ExcelReader"]
